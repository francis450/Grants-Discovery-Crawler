"""
Excel utilities for exporting grants to a shared Excel Online workbook
via the Microsoft Graph API.

Replaces the previous openpyxl / local-file approach so the crawler can
write directly to a SharePoint-hosted workbook without requiring OneDrive
local sync.

Authentication uses the MSAL client-credentials (app-only) flow.
Required Azure AD app permissions: Files.ReadWrite.All (Application).

Falls back to the legacy openpyxl local-file approach if EXCEL_OUTPUT_PATH
is configured and Graph credentials are absent.
"""

import logging
import re
from typing import List, Optional, Tuple
from urllib.parse import quote, urlparse

import requests

try:
    import msal
    MSAL_AVAILABLE = True
except ImportError:
    MSAL_AVAILABLE = False

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from config import (
    AZURE_CLIENT_ID,
    AZURE_CLIENT_SECRET,
    AZURE_TENANT_ID,
    EXCEL_OUTPUT_PATH,
    EXCEL_SHAREPOINT_URL,
    EXCEL_SHEET_NAME,
)

logger = logging.getLogger("grant_crawler")

# ── Column mapping (order matters — must match the review team's sheet) ──────
EXCEL_COLUMNS = [
    "title",
    "funding_organization",
    "grant_amount",
    "deadline",
    "deadline_review",
    "geographic_focus",
    "thematic_areas",
    "eligibility_criteria",
    "description",
    "application_url",
    "date_posted",
]

EXCEL_HEADERS = [
    "title",
    "funding_organization",
    "grant_amount",
    "deadline",
    "deadline_review",
    "geographic_focus",
    "thematic_areas",
    "eligibility_criteria",
    "description",
    "application_url",
    "date_posted",
]

GRAPH_BASE = "https://graph.microsoft.com/v1.0"
GRAPH_SCOPES = ["https://graph.microsoft.com/.default"]

# Maximum rows the Graph API accepts in a single PATCH (keep well under 10 MB)
MAX_BATCH_ROWS = 100


# ═══════════════════════════════════════════════════════════════════════════════
# Public entry point
# ═══════════════════════════════════════════════════════════════════════════════

def append_grants_to_excel(
    grants: List[dict],
    filepath: Optional[str] = None,
    sheet_name: Optional[str] = None,
) -> int:
    """
    Append new grants to the shared Excel workbook.

    Tries the Graph API path first (if credentials are configured).
    Falls back to the legacy openpyxl / local-file path if EXCEL_OUTPUT_PATH
    is set and Graph credentials are missing.

    Returns:
        Number of grants actually appended.
    """
    sheet_name = sheet_name or EXCEL_SHEET_NAME or "From Automation"

    if not grants:
        logger.info("  No grants to export to Excel.")
        return 0

    # ── Attempt 1: Microsoft Graph API ──────────────────────────────────
    if EXCEL_SHAREPOINT_URL and AZURE_TENANT_ID and AZURE_CLIENT_ID and AZURE_CLIENT_SECRET:
        if not MSAL_AVAILABLE:
            logger.warning("  msal not installed — cannot use Graph API.  pip install msal")
        else:
            try:
                return _append_via_graph(grants, sheet_name)
            except Exception as exc:
                logger.error(f"  Graph API export failed: {exc}")
                logger.info("  Falling back to local-file export…")

    # ── Attempt 2: Legacy openpyxl local file ───────────────────────────
    local_path = filepath or EXCEL_OUTPUT_PATH
    if local_path:
        return _append_via_openpyxl(grants, local_path, sheet_name)

    # ── Neither configured ──────────────────────────────────────────────
    logger.info(
        "  Excel export skipped: configure EXCEL_SHAREPOINT_URL + Azure creds, "
        "or set EXCEL_OUTPUT_PATH to a local .xlsx file."
    )
    return 0


# ═══════════════════════════════════════════════════════════════════════════════
# Graph API path
# ═══════════════════════════════════════════════════════════════════════════════

def _get_access_token() -> str:
    """Obtain an OAuth2 token via MSAL client-credentials flow."""
    authority = f"https://login.microsoftonline.com/{AZURE_TENANT_ID}"
    app = msal.ConfidentialClientApplication(
        AZURE_CLIENT_ID,
        authority=authority,
        client_credential=AZURE_CLIENT_SECRET,
    )
    result = app.acquire_token_for_client(scopes=GRAPH_SCOPES)
    if "access_token" not in result:
        error = result.get("error_description", result.get("error", "unknown"))
        raise RuntimeError(f"Failed to acquire Graph token: {error}")
    return result["access_token"]


def _parse_sharepoint_url(url: str) -> Tuple[str, str]:
    """
    Parse a SharePoint / OneDrive-for-Business URL into
    (drive_base, item_path) suitable for Graph API calls.

    Uses the site-based endpoint so we never need to guess UPNs from slugs.

    Supports:
      https://<tenant>-my.sharepoint.com/personal/<user>/Documents/<path>
      https://<tenant>.sharepoint.com/sites/<site>/Shared Documents/<path>

    Returns:
        (drive_base, item_path) where
            drive_base = Graph URL prefix for the drive
            item_path  = path inside the drive (e.g. TOH NON-PROFIT/file.xlsx)
    """
    parsed = urlparse(url)
    host = parsed.hostname or ""
    path_parts = [p for p in parsed.path.split("/") if p]

    # ── Personal OneDrive (…-my.sharepoint.com/personal/<slug>/…) ───
    if "-my.sharepoint.com" in host and "personal" in path_parts:
        idx = path_parts.index("personal")
        user_slug = path_parts[idx + 1] if idx + 1 < len(path_parts) else ""
        # Use site-based path: /sites/{host}:/personal/{slug}:/drive
        # This avoids UPN guessing entirely.
        site_path = f"/sites/{host}:/personal/{user_slug}:"
        if "Documents" in path_parts:
            doc_idx = path_parts.index("Documents")
            item_path = "/".join(path_parts[doc_idx + 1:])
        else:
            item_path = "/".join(path_parts[idx + 2:])
        return f"{site_path}/drive", item_path

    # ── SharePoint site (…/sites/<site>/Shared Documents/…) ─────────
    if "sites" in path_parts:
        idx = path_parts.index("sites")
        site_name = path_parts[idx + 1] if idx + 1 < len(path_parts) else ""
        site_path = f"/sites/{host}:/sites/{site_name}:"
        if "Shared Documents" in parsed.path:
            after = parsed.path.split("Shared Documents/", 1)[1]
        elif "Documents" in path_parts:
            doc_idx = path_parts.index("Documents")
            after = "/".join(path_parts[doc_idx + 1:])
        else:
            after = "/".join(path_parts[idx + 2:])
        return f"{site_path}/drive", after

    raise ValueError(f"Cannot parse SharePoint URL: {url}")


def _slug_to_upn(slug: str) -> str:
    """
    Convert a OneDrive personal-site slug into a UPN (email).

    Examples:
        daniela_v_dragonsino_com → daniela_v@dragonsino.com
        john_doe_contoso_com     → john.doe@contoso.com
    """
    parts = slug.split("_")
    for tld_len in (2, 1):
        if len(parts) > tld_len + 1:
            domain_parts = parts[-(tld_len + 1):]
            candidate = ".".join(domain_parts)
            user_parts = parts[:-(tld_len + 1)]
            if user_parts and domain_parts:
                return "_".join(user_parts) + "@" + candidate
    if len(parts) >= 3:
        return "_".join(parts[:-2]) + "@" + ".".join(parts[-2:])
    return slug


def _graph_headers(token: str) -> dict:
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }


def _append_via_graph(grants: List[dict], sheet_name: str) -> int:
    """Append grants to the remote Excel workbook via Graph API."""
    token = _get_access_token()
    headers = _graph_headers(token)
    drive_base, item_path = _parse_sharepoint_url(EXCEL_SHAREPOINT_URL)

    encoded_path = quote(item_path, safe="/")
    item_url = f"{GRAPH_BASE}{drive_base}/root:/{encoded_path}:"

    # ── 1. Create a persistent session ─────────────────────────────────
    session_url = f"{item_url}/workbook/createSession"
    sess_resp = requests.post(
        session_url, headers=headers,
        json={"persistChanges": True},
        timeout=30,
    )
    if sess_resp.status_code not in (200, 201):
        raise RuntimeError(
            f"createSession failed ({sess_resp.status_code}): "
            f"{sess_resp.text[:300]}"
        )
    session_id = sess_resp.json().get("id", "")
    if session_id:
        headers["Workbook-Session-Id"] = session_id

    try:
        wb_url = f"{item_url}/workbook"

        # ── 2. Ensure the target sheet exists ──────────────────────────
        sheets_resp = requests.get(
            f"{wb_url}/worksheets", headers=headers, timeout=20
        )
        sheets_resp.raise_for_status()
        sheet_names_list = [s["name"] for s in sheets_resp.json().get("value", [])]

        if sheet_name not in sheet_names_list:
            requests.post(
                f"{wb_url}/worksheets/add",
                headers=headers,
                json={"name": sheet_name},
                timeout=20,
            ).raise_for_status()
            logger.info(f"  Created new sheet: '{sheet_name}'")
            _write_row(wb_url, headers, sheet_name, 1, EXCEL_HEADERS)

        ws_url = f"{wb_url}/worksheets('{sheet_name}')"

        # ── 3. Read existing titles for deduplication ──────────────────
        existing_titles = _read_existing_titles(ws_url, headers)

        # ── 4. Build rows to append ────────────────────────────────────
        new_rows: List[List[str]] = []
        for grant in grants:
            title = grant.get("title", "")
            if title and title.strip().lower() in existing_titles:
                continue
            row = [_format_field(grant, col) for col in EXCEL_COLUMNS]
            new_rows.append(row)
            if title:
                existing_titles.add(title.strip().lower())

        if not new_rows:
            logger.info(
                f"  No new grants to add to Excel "
                f"(all {len(grants)} already exist in the sheet)."
            )
            return 0

        # ── 5. Find the next empty row ─────────────────────────────────
        used_resp = requests.get(
            f"{ws_url}/usedRange(valuesOnly=true)",
            headers=headers, timeout=20,
        )
        used_resp.raise_for_status()
        used_data = used_resp.json()
        used_addr = used_data.get("address", "")
        next_row = _next_row_from_address(used_addr)

        # ── 6. Write rows in batches ───────────────────────────────────
        total_written = 0
        for batch_start in range(0, len(new_rows), MAX_BATCH_ROWS):
            batch = new_rows[batch_start: batch_start + MAX_BATCH_ROWS]
            start_row = next_row + batch_start
            end_row = start_row + len(batch) - 1
            end_col = _col_letter(len(EXCEL_COLUMNS))
            address = f"A{start_row}:{end_col}{end_row}"
            range_url = f"{ws_url}/range(address='{address}')"
            resp = requests.patch(
                range_url, headers=headers,
                json={"values": batch}, timeout=30,
            )
            if resp.status_code not in (200, 201):
                logger.error(
                    f"  Graph API PATCH failed ({resp.status_code}): "
                    f"{resp.text[:300]}"
                )
                break
            total_written += len(batch)

        logger.info(
            f"  Appended {total_written} new grants to Excel Online "
            f"[{sheet_name}] via Graph API."
        )
        return total_written

    finally:
        # ── 7. Close the session ───────────────────────────────────────
        if session_id:
            try:
                requests.post(
                    f"{wb_url}/closeSession", headers=headers, timeout=10,
                )
            except Exception:
                pass


def _read_existing_titles(ws_url: str, headers: dict) -> set:
    """Read column A (title column) to build a dedup set."""
    titles: set = set()
    try:
        resp = requests.get(
            f"{ws_url}/range(address='A2:A5000')",
            headers=headers, timeout=20,
        )
        if resp.status_code == 200:
            values = resp.json().get("values", [])
            for row in values:
                if row and row[0]:
                    titles.add(str(row[0]).strip().lower())
    except Exception as exc:
        logger.warning(f"  Could not read existing titles for dedup: {exc}")
    return titles


def _write_row(wb_url: str, headers: dict, sheet_name: str, row_num: int, values: list):
    """Write a single row to the worksheet."""
    end_col = _col_letter(len(values))
    address = f"A{row_num}:{end_col}{row_num}"
    url = f"{wb_url}/worksheets('{sheet_name}')/range(address='{address}')"
    requests.patch(url, headers=headers, json={"values": [values]}, timeout=20)


def _next_row_from_address(address: str) -> int:
    """
    Extract the last row number from a usedRange address like
    "'From Automation'!A1:J42" and return the next row (43).
    """
    if not address:
        return 2
    match = re.search(r"(\d+)\s*$", address)
    if match:
        return int(match.group(1)) + 1
    return 2


def _col_letter(n: int) -> str:
    """Convert 1-based column number to Excel letter (1->A, 10->J, 27->AA)."""
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# Legacy openpyxl local-file path (fallback)
# ═══════════════════════════════════════════════════════════════════════════════

def _append_via_openpyxl(
    grants: List[dict], filepath: str, sheet_name: str
) -> int:
    """Original local-file implementation using openpyxl."""
    if not OPENPYXL_AVAILABLE:
        logger.warning(
            "  Excel export skipped: openpyxl not installed. "
            "Run: pip install openpyxl"
        )
        return 0

    try:
        wb = load_workbook(filepath)
    except FileNotFoundError:
        logger.error(f"  Excel file not found: {filepath}")
        return 0
    except Exception as exc:
        logger.error(f"  Error opening Excel file: {exc}")
        return 0

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        logger.info(f"  Created new sheet: '{sheet_name}'")

    existing_titles: set = set()
    last_data_row = 1  # header row
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val is not None:
            existing_titles.add(str(val).strip().lower())
            last_data_row = r

    next_row = last_data_row + 1
    if next_row == 2 and ws.cell(row=2, column=1).value is None:
        next_row = 2

    appended = 0
    for grant in grants:
        title = grant.get("title", "")
        if title and title.strip().lower() in existing_titles:
            continue

        for col_idx, field in enumerate(EXCEL_COLUMNS, start=1):
            if field == "Comments":
                ws.cell(row=next_row, column=col_idx, value="")
            else:
                ws.cell(
                    row=next_row, column=col_idx,
                    value=_format_field(grant, field),
                )
        next_row += 1
        appended += 1
        if title:
            existing_titles.add(title.strip().lower())

    if appended > 0:
        try:
            wb.save(filepath)
            logger.info(
                f"  Appended {appended} new grants to local Excel: "
                f"{filepath} [{sheet_name}]"
            )
        except PermissionError:
            logger.error(
                "  Cannot save — file may be locked by OneDrive or another user."
            )
            return 0
        except Exception as exc:
            logger.error(f"  Error saving Excel file: {exc}")
            return 0
    else:
        logger.info(
            f"  No new grants to add to Excel "
            f"(all {len(grants)} already exist in the sheet)."
        )

    wb.close()
    return appended


# ═══════════════════════════════════════════════════════════════════════════════
# Shared helpers
# ═══════════════════════════════════════════════════════════════════════════════

def sync_db_to_excel(sheet_name: Optional[str] = None) -> int:
    """
    Full sync: push every grant from the local SQLite database to the
    shared Excel workbook, skipping any titles that already exist in the
    sheet.

    This is useful as a one-time backfill (e.g. after fixing the Excel
    export) or as a periodic reconciliation command.

    Returns:
        Number of grants appended to Excel.
    """
    from utils.db_utils import get_all_grants  # local import to avoid circular

    grants = get_all_grants()
    if not grants:
        logger.info("  Database is empty — nothing to sync.")
        return 0

    logger.info(f"  Loaded {len(grants)} grants from the database.")
    return append_grants_to_excel(grants, sheet_name=sheet_name)


def _format_field(grant: dict, field: str) -> str:
    """Format a grant field for Excel output."""
    if field == "deadline_review":
        deadline = grant.get("deadline")
        if not deadline or str(deadline).strip().lower() in ("", "tbd", "null", "none"):
            return "⚠ No deadline found — verify before applying"
        return ""

    value = grant.get(field)
    if value is None:
        if field == "deadline":
            return "TBD"
        if field == "application_url":
            return "Check source page"
        return ""
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    return str(value)
